
#Comment 1: Import functions required to work with Excel. Also loads the file onto python and makes it active
from openpyxl import load_workbook
wb = load_workbook(filename='Go_Save_Me_Spreadsheet.xlsx')
sheet = wb.active

#Comment 2: Function that gets data from sheet and transfer it to python terminal for visual representation
def print_rows():
    for row in sheet.iter_rows(values_only=True):
        print(row)
    exit_select = input('Enter Y to exit')
    exit()
    
#Comment 3: Functions calculates percentages and returns which category has the highest spending
def percentage_calc():
    food_percent = sheet['E2'].value/sheet['D2'].value
    gas_percent = sheet['E4'].value/sheet['D2'].value
    other_percent = sheet['F2'].value/sheet['D2'].value
    if food_percent > gas_percent and other_percent:
        print ("{:.2%}".format(food_percent),"percent of your spending was food.\nLet's try cutting back on Starbucks lattes")
    elif gas_percent > food_percent and other_percent:
        print ("Hey buddy,","{:.2%}".format(gas_percent),"percent of your spending was gas.\nLet's save the environment by not driving so much")
    else: 
        print("{:.2%}".format(other_percent),"of your spending was other items.\nLet's cut back on the random spending, do you really need that IPhone 12?")

#Comment 4: Function adds onto the total amount of expenditures
def add_total_expen(addition):
    sheet['D2'].value += addition
    return

#Comment 5: Functions that categorizes expenditures into 3 types(Food,Gas,Other). Then adds the expenditure amount to each category total.
def expenditure_catergory(maximum_row,amount):
    add_total_expen(amount)
    category = input('Please select number for type of expense\n1.Food\n2.Gas\n3.Other\n')
    cat_location = 'C'+str(maximum_row)
    if category == '1':
        sheet[cat_location] = 'Food'
        sheet['E2'].value += amount
        return
    elif category == '2':
        sheet[cat_location] = 'Gas'
        sheet['E4'].value += amount
        return
    elif category == '3':
        sheet[cat_location] = 'Other'
        sheet['F2'].value += amount
        return
    else:
        print('Invalid Input\nPlease Try Again')
        exit()
        
#Comment 6: Functions that verifies data type of both expenditures and income
def data_verification(amount):
    try:
        verify = float(amount)
        if verify >= 0:
            return verify
        else:
            return False
    except ValueError:
        print('Invalid Entry\nPlease Try Again')
        return False

#Comment 7: Functions verifies float data type and then tracks the expenditure as a separate transaction
def add_expenditures():
    verify = input('Please Enter amount:\n')
    amount = data_verification(verify)
    if amount == False:
        exit()
    maximum_row = sheet.max_row+1
    expenditure_catergory(maximum_row,amount)
    expensiture_addition="A"+str(maximum_row)
    sheet[expensiture_addition] = amount

#Comment 8: Function verfies data for income and then adds it to the income column as well as total income. 
def add_income():
    verify = input('Please Enter amount:\n')
    amount = data_verification(verify)
    if amount == False:
        exit()
    maximum_row =sheet.max_row+1
    expensiture_addition="B"+str(maximum_row)
    sheet[expensiture_addition] = amount
    sheet['D4'].value += amount


#Comment 9:MAIN CODE where the user decides which function of the code they would like to use. Functions being expenditure tracking, income tracking, percentage report, and spreadsheet display.
if __name__ == "__main__":
    spend_or_gain = input('Please Select an Option\n1.Spend\n2.Gain\n3.Percent Report of Spending\n4.Display Sheet\n5.Exit\n')
    if spend_or_gain == '1':
        add_expenditures()
    elif spend_or_gain =='2':
        add_income()
    elif spend_or_gain =='3':
        percentage_calc()
    elif spend_or_gain =='4':
        print_rows()
    elif spend_or_gain =='5':
        exit()
    else:
        print ('Invalid Choice')
        exit()
    #wb.save(filename='Go_Save_Me_Spreadsheet.xlsx')